"""
RAG (Retrieval-Augmented Generation) API endpoints
Knowledge base search and calculator recommendations
"""

import logging
import urllib.parse
import os
import io
import json
import asyncio
from pathlib import Path
from typing import List, Optional, Dict, Any, AsyncGenerator
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Request
from fastapi.responses import JSONResponse, StreamingResponse

from app.core.security import get_optional_user, get_current_admin_user
from app.config import settings
from app.database.sqlite_models import User
from app.schemas.rag import (
    RAGSearchRequest,
    SearchResult,
    RAGSearchResponse,
    CalculatorRecommendationRequest,
    CalculatorRecommendationResponse
)
from rag.rag_pipeline import RAGPipeline

logger = logging.getLogger(__name__)

router = APIRouter()


def parse_excel_metadata(excel_content: bytes) -> Dict[str, Dict[str, Any]]:
    """
    Parse Excel spreadsheet to extract metadata for documents.

    Expected columns (case-insensitive):
    - filename: Name of the PDF file (required for matching)
    - title: Paper/document title
    - author/authors: Author names
    - journal/source: Publication source
    - year/publication_date: Publication year
    - doi: Digital Object Identifier
    - pmid: PubMed ID
    - abstract: Paper abstract
    - keywords: Keywords/tags

    Returns:
        Dictionary mapping filename (lowercase) to metadata dict
    """
    try:
        from openpyxl import load_workbook

        # Load workbook from bytes
        wb = load_workbook(filename=io.BytesIO(excel_content), read_only=True)
        ws = wb.active

        # Get header row (first row)
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')

        # Map common column name variations
        column_mapping = {
            'filename': ['filename', 'file', 'file_name', 'document', 'pdf'],
            'title': ['title', 'paper_title', 'document_title', 'name'],
            'author': ['author', 'authors', 'author(s)', 'by'],
            'journal': ['journal', 'source', 'publication', 'publisher', 'venue'],
            'year': ['year', 'publication_date', 'pub_date', 'date', 'published'],
            'doi': ['doi', 'digital_object_identifier'],
            'pmid': ['pmid', 'pubmed_id', 'pubmed'],
            'abstract': ['abstract', 'summary', 'description'],
            'keywords': ['keywords', 'tags', 'key_words'],
        }

        # Find column indices
        column_indices = {}
        for field, variations in column_mapping.items():
            for i, header in enumerate(headers):
                if header in variations:
                    column_indices[field] = i
                    break

        if 'filename' not in column_indices:
            logger.warning("Excel file missing 'filename' column - cannot match documents")
            return {}

        # Parse data rows
        metadata_map = {}
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[column_indices['filename']]:
                continue

            filename = str(row[column_indices['filename']]).strip()
            # Normalize filename for matching (lowercase, handle extensions)
            filename_key = filename.lower()
            if not filename_key.endswith('.pdf'):
                filename_key = filename_key + '.pdf'

            # Extract metadata for this file
            file_metadata = {}
            for field, col_idx in column_indices.items():
                if field != 'filename' and col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        file_metadata[field] = str(value).strip()

            if file_metadata:
                metadata_map[filename_key] = file_metadata
                logger.debug(f"Parsed metadata for {filename_key}: {list(file_metadata.keys())}")

        logger.info(f"Parsed metadata for {len(metadata_map)} documents from Excel")
        return metadata_map

    except ImportError:
        logger.error("openpyxl not installed - cannot parse Excel files")
        return {}
    except Exception as e:
        logger.error(f"Failed to parse Excel metadata: {e}")
        return {}


# Dependency injection - get RAG pipeline from app.state (initialized at startup)
def get_rag_pipeline(request: Request) -> RAGPipeline:
    """Get RAG pipeline instance from application state."""
    rag_pipeline = getattr(request.app.state, 'rag_pipeline', None)

    if rag_pipeline is None:
        logger.error("RAG pipeline not available - was it initialized at startup?")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="RAG service temporarily unavailable. Neo4j may not be connected."
        )

    return rag_pipeline


@router.post("/search", response_model=RAGSearchResponse)
async def search_knowledge_base(
    request: RAGSearchRequest,
    current_user: Optional[User] = Depends(get_optional_user),
    rag_pipeline: RAGPipeline = Depends(get_rag_pipeline)
):
    """
    Semantic search in clinical knowledge base.

    Searches across:
    - Clinical guidelines (AUA, NCCN, EAU)
    - Medical literature (PubMed articles)
    - Calculator documentation

    Uses:
    - Vector similarity search
    - Graph-augmented retrieval (includes related concepts and calculators)
    - Hybrid search (vector + keyword matching)
    """
    try:
        user_id = current_user.id if current_user else "anonymous"
        logger.info(
            f"User {user_id} searching knowledge base: {request.query[:50]}..."
        )

        # Execute RAG retrieval
        result = await rag_pipeline.retrieve_and_augment(
            query=request.query,
            k=request.limit,
            search_strategy=request.search_strategy,
            category=request.category,
            patient_context=request.patient_context
        )

        # Convert to SearchResult schema
        search_results = []
        for doc in result.documents:
            search_result = SearchResult(
                content=doc.content,
                source=doc.source,
                title=doc.title,
                relevance=doc.similarity_score,
                metadata={
                    "category": doc.category,
                    **doc.metadata
                },
                related_concepts=doc.related_concepts,
                applicable_calculators=doc.applicable_calculators
            )
            search_results.append(search_result)

        logger.info(f"Knowledge base search returned {len(search_results)} results")

        return RAGSearchResponse(
            results=search_results,
            sources=result.sources,
            metadata=result.metadata
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Knowledge base search failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Search failed: {str(e)}"
        )


@router.post("/recommend-calculators", response_model=CalculatorRecommendationResponse)
async def recommend_calculators(
    request: CalculatorRecommendationRequest,
    current_user: Optional[User] = Depends(get_optional_user),
    rag_pipeline: RAGPipeline = Depends(get_rag_pipeline)
):
    """
    Get calculator recommendations based on clinical query.

    Analyzes the clinical scenario and recommends relevant calculators
    using graph relationships and semantic similarity.
    """
    try:
        logger.info(f"User {current_user.id} requesting calculator recommendations")

        # Get calculator recommendations
        calculators = await rag_pipeline.get_calculator_recommendations(
            query=request.query,
            k=request.limit
        )

        # Enrich with calculator metadata
        from calculators.registry import registry as calc_registry

        enriched_calcs = []
        for calc_rec in calculators:
            calc_info = calc_registry.get_calculator_info(calc_rec["name"])
            if calc_info:
                enriched_calcs.append({
                    **calc_rec,
                    "calculator_id": calc_info["id"],
                    "description": calc_info["description"],
                    "category": calc_info["category"]
                })
            else:
                enriched_calcs.append(calc_rec)

        logger.info(f"Recommended {len(enriched_calcs)} calculators")

        return CalculatorRecommendationResponse(
            calculators=enriched_calcs,
            metadata={
                "query": request.query,
                "num_recommendations": len(enriched_calcs)
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Calculator recommendation failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Recommendation failed: {str(e)}"
        )


@router.get("/openevidence-query")
async def generate_openevidence_query(
    query: str,
    current_user: Optional[User] = Depends(get_optional_user)
):
    """
    Generate OpenEvidence search URL with user's credentials.

    OpenEvidence is an external evidence synthesis platform.
    This endpoint builds a search URL that the frontend can open.

    **Setup Required:**
    - User must configure OpenEvidence credentials in settings
    - Credentials are stored encrypted in SQLite
    """
    try:
        # Check if user has OpenEvidence credentials configured
        if not current_user.openevidence_username:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "OpenEvidence credentials not configured. "
                    "Please add your credentials in Settings."
                )
            )

        # Build OpenEvidence search URL
        base_url = "https://openevidence.com"
        search_url = f"{base_url}?q={urllib.parse.quote(query)}"

        logger.info(f"Generated OpenEvidence query for user {current_user.id}")

        return {
            "search_url": search_url,
            "has_credentials": True,
            "username": current_user.openevidence_username,
            "instructions": (
                "Open this URL in a new tab. You may need to log in with your "
                "OpenEvidence credentials if not already authenticated."
            )
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"OpenEvidence query generation failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate OpenEvidence query: {str(e)}"
        )


@router.get("/nsqip-link")
async def generate_nsqip_link(
    current_user: Optional[User] = Depends(get_optional_user)
):
    """
    Generate NSQIP (National Surgical Quality Improvement Program) link.

    NSQIP provides surgical risk calculators and quality metrics.

    **Access:**
    - Requires NSQIP account (typically institutional)
    - Link opens NSQIP Universal Surgical Risk Calculator
    """
    try:
        # NSQIP Universal Risk Calculator URL
        nsqip_url = "https://riskcalculator.facs.org/RiskCalculator/"

        logger.info(f"Generated NSQIP link for user {current_user.id}")

        return {
            "nsqip_url": nsqip_url,
            "calculator_name": "NSQIP Universal Surgical Risk Calculator",
            "description": (
                "Estimates surgical risk based on patient and procedure factors. "
                "Predicts 30-day morbidity and mortality."
            ),
            "instructions": (
                "Open this URL in a new tab. No login required for basic calculator. "
                "Institutional access provides additional features."
            ),
            "institution_access": current_user.institution is not None
        }

    except Exception as e:
        logger.error(f"NSQIP link generation failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate NSQIP link: {str(e)}"
        )


@router.post("/upload-documents")
async def upload_documents(
    request: Request,
    files: List[UploadFile] = File(...),
    category: str = Form(...),
    current_user: User = Depends(get_current_admin_user),
    rag_pipeline: RAGPipeline = Depends(get_rag_pipeline)
):
    """
    Upload clinical documents to build the knowledge base.

    **Admin Only** - Requires administrator role.

    Returns a streaming response (NDJSON) to keep connection alive during long processing.
    Each line is a JSON object with type: 'progress', 'heartbeat', or 'complete'.

    Supports:
    - PDF files (clinical papers, guidelines)
    - DOCX files (Word documents)
    - TXT files (plain text)
    - XLSX files (Excel spreadsheet with metadata for other documents)

    When an Excel file (.xlsx) is included in the upload:
    - It's parsed for document metadata (title, author, journal, year, doi, etc.)
    - The 'filename' column is used to match metadata to uploaded PDFs
    - Matched metadata is applied to corresponding documents during ingestion

    Documents are:
    1. Chunked into semantically meaningful segments
    2. Embedded using sentence transformers
    3. Stored in Neo4j graph database with vector index
    4. Linked to related concepts and calculators

    Category must be one of: peer_reviewed_papers, aua_guidelines, nccn_guidelines,
    aua_updates, best_practices, aua_core_curriculum, other
    """
    # Validate category first (before streaming)
    valid_categories = [
        "peer_reviewed_papers", "aua_guidelines", "nccn_guidelines",
        "aua_updates", "best_practices", "aua_core_curriculum", "other"
    ]
    if category not in valid_categories:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid category. Must be one of: {', '.join(valid_categories)}"
        )

    async def process_documents() -> AsyncGenerator[str, None]:
        """Generator that yields NDJSON lines with progress updates."""
        try:
            # Separate Excel files from document files
            allowed_doc_extensions = {'.pdf', '.docx', '.txt'}
            excel_extensions = {'.xlsx', '.xls'}

            excel_files = []
            document_files = []

            for file in files:
                file_ext = Path(file.filename).suffix.lower()
                if file_ext in excel_extensions:
                    excel_files.append(file)
                elif file_ext in allowed_doc_extensions:
                    document_files.append(file)

            # Send initial status
            yield json.dumps({
                "type": "progress",
                "message": f"Starting upload of {len(document_files)} documents",
                "total": len(document_files),
                "processed": 0
            }) + "\n"

            # Parse Excel metadata if present
            excel_metadata: Dict[str, Dict[str, Any]] = {}
            excel_parsed_count = 0

            for excel_file in excel_files:
                try:
                    excel_content = await excel_file.read()
                    parsed_metadata = parse_excel_metadata(excel_content)
                    excel_metadata.update(parsed_metadata)
                    excel_parsed_count += len(parsed_metadata)
                    logger.info(f"Parsed {len(parsed_metadata)} entries from {excel_file.filename}")
                except Exception as e:
                    logger.warning(f"Failed to parse Excel file {excel_file.filename}: {e}")

            processed_files = []
            failed_files = []

            # Create documents directory if it doesn't exist
            docs_dir = Path(settings.DOCUMENTS_DIR) if hasattr(settings, 'DOCUMENTS_DIR') else Path('./data/documents')
            docs_dir.mkdir(parents=True, exist_ok=True)

            for idx, file in enumerate(document_files):
                file_ext = Path(file.filename).suffix.lower()

                # Send heartbeat/progress before each file
                yield json.dumps({
                    "type": "progress",
                    "message": f"Processing {file.filename}",
                    "current_file": file.filename,
                    "total": len(document_files),
                    "processed": idx
                }) + "\n"

                try:
                    # Save uploaded file
                    file_path = docs_dir / f"{category}_{file.filename}"
                    file_path.parent.mkdir(parents=True, exist_ok=True)

                    content = await file.read()
                    with open(file_path, 'wb') as f:
                        f.write(content)

                    logger.info(f"Saved file: {file_path}")

                    # Build metadata
                    file_metadata = {
                        "original_filename": file.filename,
                        "uploaded_by": current_user.user_id,
                        "file_type": file_ext[1:]
                    }

                    # Check for Excel metadata match
                    filename_key = file.filename.lower()
                    if filename_key in excel_metadata:
                        matched_metadata = excel_metadata[filename_key]
                        file_metadata.update(matched_metadata)
                        logger.info(f"Applied Excel metadata to {file.filename}")

                    # Process document with periodic heartbeats
                    # Wrap ingestion in a task with background heartbeat to keep connection alive
                    yield json.dumps({
                        "type": "heartbeat",
                        "message": f"Ingesting {file.filename} (parsing, chunking, embedding)..."
                    }) + "\n"

                    # Create ingestion task and heartbeat task
                    ingestion_complete = asyncio.Event()
                    result = None
                    ingestion_error = None

                    async def do_ingestion():
                        nonlocal result, ingestion_error
                        try:
                            result = await rag_pipeline.ingest_document(
                                file_path=str(file_path),
                                category=category,
                                metadata=file_metadata
                            )
                        except Exception as e:
                            ingestion_error = e
                        finally:
                            ingestion_complete.set()

                    # Start ingestion in background
                    ingestion_task = asyncio.create_task(do_ingestion())

                    # Send heartbeats every 5 seconds while ingestion runs
                    heartbeat_count = 0
                    while not ingestion_complete.is_set():
                        try:
                            await asyncio.wait_for(ingestion_complete.wait(), timeout=5.0)
                        except asyncio.TimeoutError:
                            heartbeat_count += 1
                            yield json.dumps({
                                "type": "heartbeat",
                                "message": f"Still processing {file.filename}... ({heartbeat_count * 5}s)",
                                "current_file": file.filename,
                                "elapsed_seconds": heartbeat_count * 5
                            }) + "\n"

                    # Wait for task to fully complete
                    await ingestion_task

                    if ingestion_error:
                        raise ingestion_error

                    if result.get('status') == 'error':
                        logger.error(f"Ingestion failed for {file.filename}: {result.get('error')}")
                        failed_files.append({
                            "filename": file.filename,
                            "reason": result.get('error', 'Unknown ingestion error')
                        })
                    else:
                        processed_files.append({
                            "filename": file.filename,
                            "status": "success",
                            "chunks_created": result.get("chunks_created", 0),
                            "file_path": str(file_path),
                            "metadata_from_excel": filename_key in excel_metadata
                        })
                        logger.info(f"Successfully processed {file.filename}: {result.get('chunks_created', 0)} chunks")

                except Exception as e:
                    logger.error(f"Failed to process {file.filename}: {e}", exc_info=True)
                    failed_files.append({
                        "filename": file.filename,
                        "reason": str(e)
                    })

            # Send final result
            yield json.dumps({
                "type": "complete",
                "message": f"Processed {len(processed_files)} of {len(document_files)} documents",
                "processed": processed_files,
                "failed": failed_files,
                "category": category,
                "excel_metadata_entries": excel_parsed_count,
                "excel_files_parsed": len(excel_files)
            }) + "\n"

        except Exception as e:
            logger.error(f"Document upload failed: {e}", exc_info=True)
            yield json.dumps({
                "type": "error",
                "message": f"Upload failed: {str(e)}"
            }) + "\n"

    return StreamingResponse(
        process_documents(),
        media_type="application/x-ndjson",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Disable nginx buffering if present
        }
    )


@router.get("/system-prompt")
async def get_system_prompt(
    current_user: Optional[User] = Depends(get_optional_user)
):
    """
    Get the current urology system prompt.

    The system prompt defines the clinical note generation template and instructions
    for the LLM when generating clinic notes.

    Works without authentication (read-only endpoint).
    """
    try:
        # Read the system prompt file
        prompt_file = Path(settings.SYSTEM_PROMPT_FILE) if hasattr(settings, 'SYSTEM_PROMPT_FILE') else Path('./urology_prompt.txt')

        if not prompt_file.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="System prompt file not found"
            )

        with open(prompt_file, 'r', encoding='utf-8') as f:
            prompt_content = f.read()

        return {
            "prompt": prompt_content,
            "file_path": str(prompt_file),
            "last_modified": prompt_file.stat().st_mtime if prompt_file.exists() else None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to read system prompt: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to read system prompt: {str(e)}"
        )


@router.post("/system-prompt")
async def update_system_prompt(
    prompt: str = Form(...),
    current_user: Optional[User] = Depends(get_optional_user)
):
    """
    Update the urology system prompt.

    Only admin users can update the system prompt.
    Works without authentication for development (no admin check).

    The system prompt defines:
    - Clinical note template structure
    - Instructions for data extraction
    - Formatting guidelines
    - Clinical reasoning requirements
    """
    try:
        # Verify admin access if user is authenticated
        if current_user and current_user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only administrators can update the system prompt"
            )

        # Get prompt file path
        prompt_file = Path(settings.SYSTEM_PROMPT_FILE) if hasattr(settings, 'SYSTEM_PROMPT_FILE') else Path('./urology_prompt.txt')

        # Create backup of current prompt
        if prompt_file.exists():
            backup_file = prompt_file.with_suffix('.txt.backup')
            with open(prompt_file, 'r', encoding='utf-8') as f:
                backup_content = f.read()
            with open(backup_file, 'w', encoding='utf-8') as f:
                f.write(backup_content)
            logger.info(f"Created backup: {backup_file}")

        # Write new prompt
        with open(prompt_file, 'w', encoding='utf-8') as f:
            f.write(prompt)

        logger.info(f"System prompt updated by user {current_user.user_id}")

        return {
            "message": "System prompt updated successfully",
            "file_path": str(prompt_file),
            "backup_created": prompt_file.with_suffix('.txt.backup').exists()
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update system prompt: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update system prompt: {str(e)}"
        )


@router.get("/stats")
async def get_knowledge_base_stats(
    current_user: Optional[User] = Depends(get_optional_user),
    rag_pipeline: RAGPipeline = Depends(get_rag_pipeline)
):
    """
    Get knowledge base statistics.

    Returns:
    - Number of documents by category
    - Last update timestamp
    - Available categories
    """
    try:
        # Query Neo4j for actual statistics
        stats = await rag_pipeline.get_knowledge_base_stats()

        return stats

    except Exception as e:
        logger.error(f"Failed to get knowledge base stats: {e}", exc_info=True)
        # Return default stats when database query fails
        return {
            "total_documents": 0,
            "by_category": {
                "prostate": 0,
                "kidney": 0,
                "bladder": 0,
                "voiding": 0,
                "female": 0,
                "reconstructive": 0,
                "fertility": 0,
                "hypogonadism": 0,
                "stones": 0,
                "surgical": 0
            },
            "last_updated": None,
            "sources": ["User uploaded documents"],
            "status": "Knowledge base stats unavailable"
        }
